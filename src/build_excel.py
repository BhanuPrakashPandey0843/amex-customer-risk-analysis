"""Build executive Excel workbook from computed metrics."""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from config import EXCEL_PATH, REPORTS_DIR
from metrics_engine import load_data, load_metrics

HEADER_FILL = PatternFill("solid", fgColor="00175A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="00175A")


def _write_sheet(ws, df: pd.DataFrame, title: str):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A4"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18


def build_excel() -> Path:
    f = load_metrics()
    df = load_data()
    amex = df[df["Company"] == "American Express"]

    wb = Workbook()

    # KPI Summary
    ws = wb.active
    ws.title = "Executive KPIs"
    kpis = pd.DataFrame([
        {"KPI": "Total Industry Complaints", "Value": f.total_complaints, "Notes": "Jan 2025 – Mar 2026"},
        {"KPI": "Amex Complaints", "Value": f.amex_complaints, "Notes": f"Rank {f.amex_rank} of {f.num_companies}"},
        {"KPI": "Amex Share (%)", "Value": round(f.amex_share_pct, 1), "Notes": "Below peer average"},
        {"KPI": "YoY Growth Q1 (%)", "Value": round(f.yoy_pct, 1), "Notes": "Accelerating — early warning"},
        {"KPI": "Top 5 Issue Concentration (%)", "Value": round(f.pareto_top5_pct, 0), "Notes": "Pareto — surgical intervention possible"},
        {"KPI": "Prepaid Outlier Share (%)", "Value": round(f.prepaid_outlier_pct, 0), "Notes": "78% of industry card-usage failures"},
        {"KPI": "Timely Response (%)", "Value": round(f.amex_timely_pct, 1), "Notes": f"Rank {f.amex_timely_rank}/9"},
        {"KPI": "Closed with Explanation (%)", "Value": round(f.resolution_explanation_pct, 1), "Notes": "Competitive strength"},
        {"KPI": "Closed with Monetary Relief (%)", "Value": round(f.resolution_monetary_pct, 1), "Notes": f"Peer avg {f.peer_monetary_relief_pct:.1f}%"},
    ])
    _write_sheet(ws, kpis, "Executive KPI Summary")

    # Competitive Volume
    ws2 = wb.create_sheet("Competitive Volume")
    comp = pd.DataFrame([
        {"Rank": i, "Company": c, "Complaints": v, "Share %": round(v / f.total_complaints * 100, 1)}
        for i, (c, v) in enumerate(f.company_volume.items(), 1)
    ])
    _write_sheet(ws2, comp, "Complaint Volume by Institution")

    # Amex Issues
    ws3 = wb.create_sheet("Amex Issues")
    issues = pd.DataFrame([
        {"Rank": i, "Issue": iss, "Count": cnt, "Share %": round(cnt / f.amex_complaints * 100, 1)}
        for i, (iss, cnt) in enumerate(f.amex_issues.items(), 1)
    ])
    _write_sheet(ws3, issues, "Top Amex Complaint Issues")

    # Issue YoY
    ws4 = wb.create_sheet("Issue YoY Trends")
    yoy_df = pd.DataFrame([
        {"Issue": k, "Q1 2025": v["q1_2025"], "Q1 2026": v["q1_2026"], "YoY %": v["yoy_pct"]}
        for k, v in f.issue_yoy_changes.items()
    ]).sort_values("YoY %", ascending=False)
    _write_sheet(ws4, yoy_df, "Issue YoY Change (Q1 2025 vs Q1 2026)")

    # Geographic
    ws5 = wb.create_sheet("Geography")
    geo = pd.DataFrame([
        {"State": s, "Complaints": v, "Share %": round(v / f.amex_complaints * 100, 1)}
        for s, v in f.top_states.items()
    ])
    _write_sheet(ws5, geo, "Top States — Amex Complaints")

    # Risks
    ws6 = wb.create_sheet("Risk Register")
    risks = pd.DataFrame(f.risks)
    _write_sheet(ws6, risks, "Enterprise Risk Register")

    # Recommendations
    ws7 = wb.create_sheet("Recommendations")
    recs = pd.DataFrame(f.recommendations)
    _write_sheet(ws7, recs, "Strategic Recommendations")

    # Pivot-style: Product x Issue for Amex
    ws8 = wb.create_sheet("Product Issue Matrix")
    pivot = pd.crosstab(amex["Product"], amex["Issue"])
    top_products = amex["Product"].value_counts().head(5).index
    top_issues = amex["Issue"].value_counts().head(8).index
    pivot_subset = pivot.loc[top_products, top_issues].reset_index()
    _write_sheet(ws8, pivot_subset, "Product × Issue Cross-Tab (Top 5 × Top 8)")

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    return EXCEL_PATH


if __name__ == "__main__":
    path = build_excel()
    print(f"Excel saved: {path}")
